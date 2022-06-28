# Utils Module

import os
import openpyxl as xl
import win32com.client as win32


def getCell(sheet: xl.worksheet, targetText):
    """
    targetText 에 해당하는 cell 객체를 반환.

    Args:
        sheet (xl.worksheet): openpyxl.worksheet 객체
        targetText (string): cell에 적힌 문자열

    Returns:
        openpyxl.cell: cell 객체
    """  

    for row_data in sheet.iter_rows():
        for cell in row_data:
            if cell.value == targetText:
                return cell


def getHeaderRowIdx(sheet: xl.worksheet, headerText):
    return getCell(sheet, headerText).row


def getHeaderColIdx(sheet: xl.worksheet, headerText):
    return getCell(sheet, headerText).column


def deleteDuplicatedColumn(sheet: xl.worksheet, exceptColList):
    exceptColList.sort()
    
    for colStr in exceptColList:
        sheet.delete_cols(xl.utils.cell.column_index_from_string(colStr))


def getContentsRowIdx(sheet: xl.worksheet, headerRowIdx: int):
    for row_data in sheet.iter_rows(min_row = headerRowIdx):	# min_row는 시작 행을 지정
        for cell in row_data:
            # Contents 시작 행 찾는 부분
            increaseRow = 1
            
            # Merge된 Cell이 아닌 Column이 처음으로 나오는 Cell을 찾음.
            while type(cell.offset(increaseRow, 0)).__name__ == 'MergedCell':
                increaseRow += 1
        
            return cell.row + increaseRow


def getSheetContents(sheet: xl.worksheet, headerRowIdx: int, contentsRowIdx: int, headerList, exceptContentsWordList):
    """
    인자로 받은 정보를 기반으로, Contents 내용을 담은 dictionary를 반환.

    Args:
        sheet (xl.worksheet): openpyxl.worksheet 객체
        headerRowIdx (int): header 시작 행 index
        contentsRowIdx (int): contents 시작 행 index
        headerList (list): 읽어야할 header 리스트
        exceptContentsWordList (list): 제외해야할 contents 문자열 리스트

    Returns:
        dictionary: contents (ex. { rowIdx :contentList }) ) 를 담은 객체
    """    
    headerColIdxList = []
    # { rowIdx :contentList }
    resultDic = {}
    # MergedCell 고려.
    headerRows = sheet[headerRowIdx : contentsRowIdx - 1]
    header_row = ()
    
    if len(headerRows) > 1:
        for headerRow in headerRows:
            header_row += headerRow
    else:
        header_row = headerRow
    
    # 수집하려는 Header Column Index 찾는 부분
    for cell in header_row:
        if cell.value in headerList:
            # print('[', cell.value, ']')
            headerColIdxList.append(getHeaderColIdx(sheet, cell.value))
    
    if len(headerColIdxList) > 1:
             for row_data in sheet.iter_rows(min_row = contentsRowIdx):	# min_row는 시작 행을 지정
                contentList = []
                for cell in row_data:
                    if cell.column in headerColIdxList:
                        if cell.value != None:
                            contentList.append(cell.value)
                            # print('[', cell.value, ']')
                resultDic[cell.row] = contentList
                
    if len(resultDic) > 1:
    # 수집된 Row 중, 제외해야 하는 단어가 포함된 Row를 제거한다.
        delItemList = []
        item = resultDic.items()
        
        for key, value in item:
            # filter(lambda x: x in value[1], exceptContentsWordList)
            if len([word for word in exceptContentsWordList if word in value[1]]) > 0:
                delItemList.append(key)
                
    for key in delItemList:
        del resultDic[key]
        
    return resultDic


def getMostRecentFile(folder_path):

    # each_file_path_and_gen_time: 각 file의 경로와, 생성 시간을 저장함
    each_file_path_and_gen_time = []
    for each_file_name in os.listdir(folder_path):
        # getctime: 입력받은 경로에 대한 생성 시간을 리턴
        each_file_path = folder_path + '\\' +  each_file_name
        each_file_gen_time = os.path.getctime(each_file_path)
        each_file_path_and_gen_time.append(
            (os.path.abspath(each_file_path), each_file_gen_time)
        )

    # 가장 생성시각이 큰(가장 최근인) 파일을 리턴 
    return max(each_file_path_and_gen_time, key=lambda x: x[1])[0]


def convertXlsToXlsx(xls_file):
    """
    win32com api를 사용하여 xls 파일을 실행 후 xlsx 파일로 저장한다.
    """
    # excel 파일을 다룰 수 있는 윈도우 프로그램 실행
    # excel_app = win32.gencache.EnsureDispatch('Excel.Application')
    excel_app = win32.Dispatch("Excel.Application")
    excel_app.Visible = True
    
    wb = excel_app.Workbooks.Open(xls_file)
    # sheet = wb.Worksheets('이용대금명세서_2206(신용.체크)_20220616103')
    sheet = wb.ActiveSheet
    
    wb.SaveAs(xls_file + 'x', FileFormat = 51)  # FileFormat = 51은 .xlsx 확장자, 56은 .xls
    wb.Close()
    excel_app.Application.Quit()